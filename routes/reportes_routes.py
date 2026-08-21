from flask import Blueprint, request, jsonify, send_file
import io
import os
import re
import pandas as pd
import traceback
from datetime import datetime, date
from sqlalchemy import func 
from extensions import db

# MODELOS
from models.pago import Pago
from models.alumno import Alumno
from models.estructura_pago import EstructuraPago
from models.grupo import Grupo 
from models.abonos_pago import AbonoPago

from helpers import registrar_accion, obtener_id_admin
import subprocess

# IMPORTAMOS ZONEINFO PARA FORZAR HORA DE MEXICO EN TODO EL ARCHIVO
from zoneinfo import ZoneInfo 

# IMPORTAMOS EL ESCUDO DE SEGURIDAD
from flask_jwt_extended import verify_jwt_in_request, get_jwt

# EXCEL PRO
from openpyxl.styles import Font, PatternFill, Alignment

# PDF PRO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT 

reportes_bp = Blueprint('reportes_bp', __name__)

# RUTA DEL LOGO
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
RUTA_LOGO = os.path.join(BASE_DIR, 'static', 'img', 'LOGO3.png')

print(f"\n--- REVISANDO LOGO EN: {RUTA_LOGO} ---")
if os.path.exists(RUTA_LOGO):
    print(" LOGO ENCONTRADO\n")
else:
    print(" LOGO NO ENCONTRADO - Revisa la carpeta static/img\n")

def limpiar_param(val):
    if val in [None, "null", "undefined", "", "Todos", "0"]: return None
    return val

# FUNCIONES MAESTRAS PARA OBTENER LA HORA DE MEXICO
def ahora_mx():
    return datetime.now(ZoneInfo("America/Mexico_City"))

def hoy_mx():
    return ahora_mx().date()

# 🔥 LA FUNCIÓN DEFINITIVA: ORDEN CURRICULAR EXACTO + BLOQUES + DICCIONARIO 🔥
def ordenar_conceptos_escolares(lista_adeudos):
    meses_orden = {
        'AGOSTO': 1, 'SEPTIEMBRE': 2, 'OCTUBRE': 3, 'NOVIEMBRE': 4,
        'DICIEMBRE': 5, 'ENERO': 6, 'FEBRERO': 7, 'MARZO': 8,
        'ABRIL': 9, 'MAYO': 10, 'JUNIO': 11, 'JULIO': 12
    }

    # 🧠 LISTAS MAESTRAS DE ORDEN CURRICULAR POR SEMESTRE (En orden de tus fotos)
    MATERIAS_POR_SEMESTRE = {
        1: ['LA MATERIA Y SUS INTERACCIONES', 'CIENCIAS SOCIALES I', 'CULTURA DIGITAL I', 'PENSAMIENTO MATEMATICO I', 'PENSAMIENTO MATEMÁTICO I', 'LENGUA Y COMUNICACION I', 'LENGUA Y COMUNICACIÓN I', 'INGLES I', 'INGLÉS I', 'HUMANIDADES I', 'LABORATORIO DE INV', 'LABORATORIO DE INVESTIGACION', 'CIENCIAS NATURALES, EXPERIMENTALES Y TECNOLOGIA I', 'PENSAMIENTO FILOSOFICO Y HUMANIDADES I', 'EDUCACION SOCIOEMOCIONAL I', 'ARTISTICA', 'ARTÍSTICA', 'EDU FISICA', 'EDUCACION FISICA', 'HABILIDAD'],
        2: ['CONSERVACION DE LA ENERGIA', 'CONSERVACIÓN DE LA ENERGÍA', 'CIENCIAS SOCIALES II', 'CULTURA DIGITAL II', 'PENSAMIENTO MATEMATICO II', 'PENSAMIENTO MATEMÁTICO II', 'LENGUA Y COMUNICACION II', 'LENGUA Y COMUNICACIÓN II', 'INGLES II', 'INGLÉS II', 'HUMANIDADES II', 'TALLER DE CIENCIAS I', 'CIENCIAS NATURALES, EXPERIMENTALES Y TECNOLOGIA II', 'PENSAMIENTO FILOSOFICO Y HUMANIDADES II', 'EDUCACION SOCIOEMOCIONAL II', 'SOCIOLOGIA', 'SOCIOLOGÍA', 'PSICOLOGIA', 'PSICOLOGÍA'],
        3: ['ECOSISTEMAS', 'PENSAMIENTO MATEMATICO III', 'PENSAMIENTO MATEMÁTICO III', 'LENGUA Y COMUNICACION III', 'LENGUA Y COMUNICACIÓN III', 'INGLES III', 'INGLÉS III', 'HUMANIDADES III', 'TALLER DE CIENCIAS II', 'CONCEPTOS DE DESARROLLO', 'CONCEPTOS DE LA COMUNIDAD', 'CIENCIAS NATURALES, EXPERIMENTALES Y TECNOLOGIA III', 'PENSAMIENTO FILOSOFICO Y HUMANIDADES III', 'EDUCACION SOCIOEMOCIONAL III', 'INFORMATICA III', 'INFORMÁTICA III', 'HISTORIA DEL ARTE', 'PEDAGOGIA', 'PEDAGOGÍA'],
        4: ['REACCIONES QUIM', 'REACCIONES QUÍM', 'REACIONES QUIM', 'REACCIONES QUIMICAS', 'CONCIENCIA HISTORICA I', 'CONCIENCIA HISTÓRICA I', 'TALLER DE CULTURA DIGITAL', 'TEMAS SELECTOS DE MATEMATICAS I', 'TEMAS SELECTOS DE MATEMÁTICAS I', 'PENSAMIENTO LITERARIO', 'INGLES IV', 'INGLÉS IV', 'ESPACIO Y SOCIEDAD', 'CIENCIAS SOCIALES III', 'DIAGNOSTICO COMUNITARIO', 'DIAGNÓSTICO COMUNITARIO', 'CIENCIAS NATURALES, EXPERIMENTALES Y TECNOLOGIA IV', 'PENSAMIENTO MATEMATICO IV', 'PENSAMIENTO MATEMÁTICO IV', 'EDUCACION SOCIOEMOCIONAL IV', 'DERECHO I', 'ETIMOLOGIAS', 'ETIMOLOGÍAS'],
        5: ['ENERGIA EN LOS PROCESOS', 'ENERGÍA EN LOS PROCESOS', 'CONCIENCIA HISTORICA II', 'CONCIENCIA HISTÓRICA II', 'CALCULO DIFERENCIAL', 'CÁLCULO DIFERENCIAL', 'TEMAS SELECTOS DE BIOLOGIA I', 'TEMAS SELECTOS DE BIOLOGÍA I', 'TEMAS SELECTOS DE QUIMICA I', 'TEMAS SELECTOS DE QUÍMICA I', 'CONTABILIDAD I', 'FORMAS LEGALES', 'FORMULACION DE PROYECTOS', 'FORMULACIÓN DE PROYECTOS', 'CIENCIAS NATURALES, EXPERIMENTALES Y TECNOLOGIA V', 'PENSAMIENTO MATEMATICO V', 'PENSAMIENTO MATEMÁTICO V', 'INGLES V', 'INGLÉS V', 'EDUCACION SOCIOEMOCIONAL V', 'INFORMATICA V', 'INFORMÁTICA V', 'MAT FINANCIERAS', 'MAT. FINANCIERAS', 'FILOSOFIA', 'FILOSOFÍA'],
        6: ['ORGANISMOS', 'CONCIENCIA HISTORICA III', 'CONCIENCIA HISTÓRICA III', 'TEMAS SELECTOS DE MATEMATICAS II', 'TEMAS SELECTOS DE MATEMÁTICAS II', 'CALCULO INTEGRAL', 'CÁLCULO INTEGRAL', 'TEMAS SELECTOS DE BIOLOGIA II', 'TEMAS SELECTOS DE BIOLOGÍA II', 'TEMAS SELECTOS DE QUIMICA II', 'TEMAS SELECTOS DE QUÍMICA II', 'CONTABILIDAD II', 'SOCIEDADES MERCANTILES', 'INSTRUMENTOS DE PROYECTOS', 'CIENCIAS NATURALES, EXPERIMENTALES Y TECNOLOGIA VI', 'CULTURA DIGITAL III', 'PENSAMIENTO MATEMATICO VI', 'PENSAMIENTO MATEMÁTICO VI', 'EDUCACION SOCIOEMOCIONAL VI', 'INGLES VI', 'INGLÉS VI', 'INFORMATICA VI', 'INFORMÁTICA VI', 'ADMINISTRACION', 'ADMINISTRACIÓN']
    }

    # 🔥 NUEVA MAGIA: Averiguar la posición exacta de la materia en tu currícula
    def obtener_indice_materia(semestre, concepto):
        if semestre in MATERIAS_POR_SEMESTRE:
            for idx, materia in enumerate(MATERIAS_POR_SEMESTRE[semestre]):
                if re.search(rf'\b{re.escape(materia)}\b', concepto):
                    return idx
        return 999 # Si es una materia rara que no está en lista, la manda al fondo

    def rescatar_semestre(concepto):
        # Busca la materia en todas las listas para adivinar el semestre
        for sem, materias in MATERIAS_POR_SEMESTRE.items():
            for materia in materias:
                if re.search(rf'\b{re.escape(materia)}\b', concepto):
                    return sem
        
        # Si no la encuentra, intenta adivinar por el número romano (I, II, III...)
        match_roman = re.search(r'\b(I|II|III|IV|V|VI)\b$', concepto.replace('.', '').strip())
        if match_roman:
            return {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5, 'VI': 6}[match_roman.group(1)]
        
        return None

    # ========================================================
    # 1. ASIGNAR SEMESTRE Y ETIQUETA
    # ========================================================
    for item in lista_adeudos:
        concepto_original = str(item.get('concepto', ''))
        concepto_upper = concepto_original.upper().strip()
        
        es_mes = any(mes in concepto_upper for mes in meses_orden)
        es_inscripcion = 'INSCRIPCIÓN' in concepto_upper or 'INSCRIPCION' in concepto_upper
        es_legalizacion = 'LEGALIZACION' in concepto_upper or 'LEGALIZACIÓN' in concepto_upper or 'CERTIFICADO' in concepto_upper
        
        semestre_bd = item.get('semestre')
        
        semestre_final = semestre_bd
        if not semestre_final:
            match_sem = re.search(r'(\d+)°\s*SEM', concepto_upper)
            if match_sem:
                semestre_final = int(match_sem.group(1))
            else:
                semestre_final = rescatar_semestre(concepto_upper)
        
        item['semestre_calculado'] = semestre_final or 99 

        if not es_mes and not es_inscripcion and not es_legalizacion:
            if "SEM" not in concepto_upper and semestre_final:
                item['concepto'] = f"{concepto_original} - {semestre_final}° SEM"

    # ========================================================
    # 2. ORDENAR ESTRICTAMENTE
    # ========================================================
    def criterio_orden(item):
        concepto = str(item.get('concepto', '')).upper().strip()
        semestre_real = item.get('semestre_calculado', 99)
        
        # BLOQUE 3: LEGALIZACIÓN
        if 'LEGALIZACION' in concepto or 'LEGALIZACIÓN' in concepto or 'CERTIFICADO' in concepto:
            return (3, 999, 999, 999, 999, concepto) 
        
        es_inscripcion = 'INSCRIPCIÓN' in concepto or 'INSCRIPCION' in concepto
        es_mes_val = next((mes for mes in meses_orden if mes in concepto), None)

        # BLOQUE 1: INSCRIPCIONES Y MESES
        if es_inscripcion:
            return (1, semestre_real, 1, 0, 0, concepto)
        elif es_mes_val:
            return (1, semestre_real, 2, meses_orden[es_mes_val], 0, concepto)
        
        # BLOQUE 2: EXTRAS ORDENADOS POR CURRÍCULA
        else:
            if 'REG.OPORTUNIDAD' in concepto or 'REG OPORTUNIDAD' in concepto or 'REG. OPORTUNIDAD' in concepto:
                sub_prioridad = 1
            elif 'PRIMERA OPORTUNIDAD' in concepto or '1RA OPORTUNIDAD' in concepto or '1RA. OPORTUNIDAD' in concepto:
                sub_prioridad = 2
            elif 'SEGUNDA OPORTUNIDAD' in concepto or '2DA OPORTUNIDAD' in concepto or '2DA. OPORTUNIDAD' in concepto:
                sub_prioridad = 3
            elif 'TERCERA OPORTUNIDAD' in concepto or '3RA OPORTUNIDAD' in concepto or '3RA. OPORTUNIDAD' in concepto:
                sub_prioridad = 4
            elif 'OPORTUNIDAD' in concepto:
                sub_prioridad = 5
            else:
                sub_prioridad = 6 # Materia regular
            
            # Buscamos en qué posición está en la currícula oficial (1ra, 2da, 3ra...)
            orden_materia = obtener_indice_materia(semestre_real, concepto)

            return (2, semestre_real, 3, sub_prioridad, orden_materia, concepto)

    return sorted(lista_adeudos, key=criterio_orden)

# --- 1. PIE DE PÁGINA ---
def pie_de_pagina(canvas, doc):
    canvas.saveState()
    estilos = getSampleStyleSheet()
    texto = f"SISTEMA DE CONTROL DE PAGOS  —  Página {doc.page}"
    p = Paragraph(f"<font color='grey' size=8>{texto}</font>", estilos['Normal'])
    p.wrapOn(canvas, letter[0], 50)
    p.drawOn(canvas, letter[0]/2 - 80, 20)
    canvas.restoreState()

# --- 2. ENCABEZADO OFICIAL ---
def crear_encabezado_logo(elementos, titulo_reporte, estilos):
    try:
        if os.path.exists(RUTA_LOGO):
            img = Image(RUTA_LOGO, width=65, height=65) 
        else:
            img = ""
    except: img = ""

    estilo_escuela = estilos['Title']
    estilo_escuela.fontSize = 17
    estilo_escuela.alignment = TA_LEFT
    estilo_escuela.leftIndent = 0 
    estilo_escuela.textColor = colors.HexColor("#0f172a")
    estilo_escuela.spaceAfter = 2

    estilo_sistema = estilos['Heading2']
    estilo_sistema.fontSize = 11
    estilo_sistema.alignment = TA_LEFT
    estilo_sistema.leftIndent = 0
    estilo_sistema.textColor = colors.HexColor("#64748b")
    estilo_sistema.spaceAfter = 6

    estilo_tipo_reporte = estilos['Normal']
    estilo_tipo_reporte.fontSize = 12
    estilo_tipo_reporte.fontName = 'Helvetica-Bold'
    estilo_tipo_reporte.alignment = TA_LEFT
    estilo_tipo_reporte.leftIndent = 0
    estilo_tipo_reporte.textColor = colors.HexColor("#1d4ed8") 
    estilo_tipo_reporte.spaceAfter = 2

    estilo_fecha = estilos['Normal']
    estilo_fecha.fontSize = 9
    estilo_fecha.leftIndent = 0
    estilo_fecha.textColor = colors.grey

    data = [[img, [
        Paragraph("<b>EPC JUAN MIRANDA URESTI N°16</b>", estilo_escuela),
        Paragraph("SISTEMA DE CONTROL DE PAGOS", estilo_sistema),
        Paragraph(f"{titulo_reporte}", estilo_tipo_reporte),
        Paragraph(f"Fecha de reporte: {ahora_mx().strftime('%d/%m/%Y %H:%M')}", estilo_fecha)
    ]]]

    t = Table(data, colWidths=[70, 460])
    t.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (0,0), 0), 
        ('LINEBELOW', (0,0), (-1,-1), 1.5, colors.HexColor("#1E293B")),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    
    elementos.append(t)
    elementos.append(Spacer(1, 15))

# --- 3. MOTOR PDF DEUDORES ---
def generar_pdf_deudores_pro(data_agrupada, titulo_str, nombre_archivo, mensaje_vacio="No hay registros.", descargar=True, es_individual=False):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, title=titulo_str,
                            rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    estilos = getSampleStyleSheet()
    elementos = []
    crear_encabezado_logo(elementos, titulo_str, estilos)
    
    if not data_agrupada:
        estilo_aviso = estilos['Normal']
        estilo_aviso.fontSize = 12
        estilo_aviso.textColor = colors.HexColor("#475569")
        elementos.append(Paragraph(f"<i>{mensaje_vacio}</i>", estilo_aviso))
    else:
        total_general = 0
        estilo_nombre = estilos['Normal']
        estilo_nombre.fontSize = 11
        estilo_nombre.leftIndent = 0

        for alu in data_agrupada:
            nombre_str = f"<font color='#0f172a'><b>{alu['Alumno']}</b></font> — <font color='#b91c1c'><b>{alu['Matrícula']}</b></font>"
            p_nombre = Paragraph(nombre_str, estilo_nombre)
            espacio = Spacer(1, 5)
            
            tabla_data = [["Concepto", "Monto Restante"]] 
            for c in alu["Conceptos"]:
                tabla_data.append([c["concepto"], "${:,.2f}".format(c["monto"])])
            tabla_data.append(["TOTAL ADEUDO", "${:,.2f}".format(alu['Total'])])
            total_general += alu["Total"]
            
            # Ajuste de ancho de columnas para que el titulo "Monto Restante" quepa perfectamente
            t = Table(tabla_data, colWidths=[430, 100], hAlign='LEFT', repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#b91c1c")), 
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
                ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BACKGROUND', (0,-1), (-1,-1), colors.whitesmoke),
                ('TEXTCOLOR', (0,-1), (-1,-1), colors.HexColor("#b91c1c")),
            ]))
            
            if es_individual:
                elementos.append(p_nombre)
                elementos.append(espacio)
                elementos.append(t)
                elementos.append(Spacer(1, 15))
            else:
                bloque = []
                bloque.append(p_nombre)
                bloque.append(espacio)
                bloque.append(t)
                bloque.append(Spacer(1, 15))
                elementos.append(KeepTogether(bloque))
            
        elementos.append(Spacer(1, 20))
        estilo_total = estilos['Normal']
        estilo_total.fontSize = 12
        elementos.append(Paragraph(f"<b>TOTAL GENERAL DE ADEUDOS: ${total_general:,.2f}</b>", estilo_total))
    
    doc.build(elementos, onFirstPage=pie_de_pagina, onLaterPages=pie_de_pagina)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='application/pdf', as_attachment=descargar, download_name=nombre_archivo)

# --- 4. MOTOR PDF GENÉRICO CON TOTALIZADOR ---
def generar_pdf_generico_pro(titulo_str, data, color_tema, nombre_archivo, total_suma=None, titulo_total="TOTAL", mensaje_vacio="No se encontraron registros."):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, title=titulo_str,
                            rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    estilos = getSampleStyleSheet()
    elementos = []
    crear_encabezado_logo(elementos, titulo_str, estilos)
    
    if not data:
        estilo_aviso = estilos['Normal']
        estilo_aviso.fontSize = 12
        estilo_aviso.textColor = colors.HexColor("#475569")
        elementos.append(Paragraph(f"<i>{mensaje_vacio}</i>", estilo_aviso))
    else:
        headers = list(data[0].keys())
        cuerpo = [headers]
        for d in data: cuerpo.append([str(valor) for valor in d.values()])
        
        anchos = [530 / len(headers)] * len(headers)
        if "Folio" in headers and "Concepto" in headers:
            anchos = [60, 200, 200, 70] 
        elif "Fecha" in headers and "Monto" in headers:
            anchos = [70, 195, 195, 70] 
        elif "Matrícula" in headers and "Estatus" in headers:
            anchos = [80, 350, 100]    

        t = Table(cuerpo, colWidths=anchos, hAlign='LEFT') 
        
        estilo_tabla = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), color_tema),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")), 
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('TEXTCOLOR', (0,1), (-1,-1), colors.black), 
        ])
        
        for i in range(1, len(cuerpo)):
            bg_color = colors.HexColor("#f8fafc") if i % 2 == 0 else colors.white
            estilo_tabla.add('BACKGROUND', (0,i), (-1,i), bg_color)

        t.setStyle(estilo_tabla)
        elementos.append(t)
        
        if total_suma is not None:
            elementos.append(Spacer(1, 20))
            estilo_total = estilos['Normal']
            estilo_total.fontSize = 12
            elementos.append(Paragraph(f"<b>{titulo_total}: ${total_suma:,.2f}</b>", estilo_total))
        
    doc.build(elementos, onFirstPage=pie_de_pagina, onLaterPages=pie_de_pagina)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=nombre_archivo)

# --- 5. EXCEL PRO ---
def generar_excel(data, nombre_hoja, nombre_archivo, mensaje_vacio="No hay información para mostrar"):
    df = pd.DataFrame(data) if data else pd.DataFrame([{"Aviso": mensaje_vacio}])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=nombre_hoja)
        ws = writer.sheets[nombre_hoja]
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(horizontal="center")
            max_len = max(df[column].astype(str).map(len).max(), len(str(column))) + 4
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len
        ws.auto_filter.ref = ws.dimensions
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=nombre_archivo)

# ==========================================
#                ENDPOINTS
# ==========================================

@reportes_bp.route("/reportes/dashboard", methods=["GET", "OPTIONS"])
def dashboard_ingresos():
    if request.method == "OPTIONS": return jsonify({}), 200
    
    try: verify_jwt_in_request()
    except Exception: return jsonify({"error": "Sesión inválida o token ausente"}), 401
    
    operador = get_jwt()
    if operador.get('rol') not in ['SISTEMAS', 'ADMIN', 'LECTURA']:
        return jsonify({"error": "Acceso denegado"}), 403

    try:
        hoy_str = hoy_mx().strftime("%Y-%m-%d")
        inicio_hoy = f"{hoy_str} 00:00:00"
        fin_hoy = f"{hoy_str} 23:59:59"
        
        p_mes_str = hoy_mx().replace(day=1).strftime("%Y-%m-%d")
        inicio_mes = f"{p_mes_str} 00:00:00"

        suma_hoy = db.session.query(func.sum(AbonoPago.monto_abono))\
            .filter(AbonoPago.fecha_abono.between(inicio_hoy, fin_hoy)).scalar() or 0.0
            
        suma_mes = db.session.query(func.sum(AbonoPago.monto_abono))\
            .filter(AbonoPago.fecha_abono.between(inicio_mes, fin_hoy)).scalar() or 0.0

        return jsonify({"ingresos_hoy": float(suma_hoy), "ingresos_mes": float(suma_mes)})
    except Exception as e: 
        print("ERROR DASHBOARD:", str(e))
        return jsonify({"error": "Dashboard error"}), 500


@reportes_bp.route("/reportes/corte-caja", methods=["GET", "OPTIONS"])
def corte_caja():
    if request.method == "OPTIONS": return jsonify({}), 200

    try: verify_jwt_in_request()
    except Exception: return jsonify({"error": "Sesión inválida o token ausente"}), 401
    
    operador = get_jwt()
    if operador.get('rol') not in ['SISTEMAS', 'ADMIN']:
        return jsonify({"error": "Solo administradores pueden descargar el corte de caja"}), 403

    try:
        formato = request.args.get('formato', 'pdf')
        
        hoy_str = hoy_mx().strftime("%Y-%m-%d")
        inicio_hoy = f"{hoy_str} 00:00:00"
        fin_hoy = f"{hoy_str} 23:59:59"
        
        abonos = db.session.query(AbonoPago, Pago, Alumno, EstructuraPago)\
            .join(Pago, AbonoPago.id_pago == Pago.id_pago)\
            .join(Alumno, Pago.id_alumno == Alumno.id_alumno)\
            .join(EstructuraPago, Pago.id_estructura == EstructuraPago.id_estructura)\
            .filter(AbonoPago.fecha_abono.between(inicio_hoy, fin_hoy))\
            .order_by(AbonoPago.id_abono.asc())\
            .all()
        
        data = []
        total_suma = 0 
        for abono, p, a, e in abonos:
            monto_real = float(abono.monto_abono)
            total_suma += monto_real
            data.append({
                "Folio": abono.folio, 
                "Alumno": f"{a.apellido} {a.nombre}".upper(), 
                "Concepto": e.concepto, 
                "Monto": "${:,.2f}".format(monto_real)
            })
        
        try: registrar_accion(operador.get('id'), "DESCARGA_CORTE_CAJA", f"Descargó el Corte de Caja del día de hoy en formato {formato.upper()}")
        except: pass

        mensaje_vacio = "No existen ingresos registrados para el día de hoy." 

        if formato == 'excel': 
            if data: data.append({"Folio": "", "Alumno": "", "Concepto": "TOTAL DEL DÍA", "Monto": "${:,.2f}".format(total_suma)})
            return generar_excel(data, "Corte", f"Corte_{hoy_str}.xlsx", mensaje_vacio)
            
        return generar_pdf_generico_pro(f"CORTE DE CAJA DIARIO", data, colors.HexColor("#1E293B"), "Corte.pdf", total_suma=total_suma, titulo_total="TOTAL INGRESOS DEL DÍA", mensaje_vacio=mensaje_vacio)
    except Exception as e: return jsonify({"error": str(e)}), 500
    
@reportes_bp.route("/reportes/ingresos", methods=["GET", "OPTIONS"])
def reporte_ingresos():
    if request.method == "OPTIONS": return jsonify({}), 200

    try: verify_jwt_in_request()
    except Exception: return jsonify({"error": "Sesión inválida o token ausente"}), 401
    
    operador = get_jwt()
    if operador.get('rol') not in ['SISTEMAS', 'ADMIN']:
        return jsonify({"error": "No tienes permisos para descargar este reporte"}), 403

    try:
        inicio = request.args.get('inicio')
        fin = request.args.get('fin')
        formato = request.args.get('formato', 'pdf')
        
        try: registrar_accion(operador.get('id'), "DESCARGA_INGRESOS", f"Descargó histórico de ingresos ({inicio} a {fin}) en formato {formato.upper()}")
        except: pass

        inicio_completo = f"{inicio} 00:00:00"
        fin_completo = f"{fin} 23:59:59"

        abonos = db.session.query(AbonoPago, Pago, Alumno, EstructuraPago)\
            .join(Pago, AbonoPago.id_pago == Pago.id_pago)\
            .join(Alumno, Pago.id_alumno == Alumno.id_alumno)\
            .join(EstructuraPago, Pago.id_estructura == EstructuraPago.id_estructura)\
            .filter(AbonoPago.fecha_abono.between(inicio_completo, fin_completo))\
            .order_by(AbonoPago.fecha_abono.asc(), AbonoPago.id_abono.asc())\
            .all()
        
        data = []
        total_suma = 0 
        for abono, p, a, e in abonos:
            monto_real = float(abono.monto_abono)
            total_suma += monto_real
            data.append({
                "Fecha": abono.fecha_abono.strftime("%d/%m/%Y"), 
                "Alumno": f"{a.apellido} {a.nombre}".upper(), 
                "Concepto": e.concepto, 
                "Monto": "${:,.2f}".format(monto_real)
            })
        
        mensaje_vacio = f"No existen ingresos registrados en el periodo del {inicio} al {fin}." 

        if formato == 'excel': 
            if data: data.append({"Fecha": "", "Alumno": "", "Concepto": "TOTAL INGRESOS", "Monto": "${:,.2f}".format(total_suma)})
            return generar_excel(data, "Ingresos", "Historico_Ingresos.xlsx", mensaje_vacio)
            
        return generar_pdf_generico_pro(f"HISTORIAL DE INGRESOS ({inicio} al {fin})", data, colors.HexColor("#1E293B"), "Historico.pdf", total_suma=total_suma, titulo_total="TOTAL GENERAL DE INGRESOS", mensaje_vacio=mensaje_vacio)
    except Exception as e: return jsonify({"error": str(e)}), 500

@reportes_bp.route("/reportes/deudores", methods=["GET", "OPTIONS"])
def reporte_deudores():
    if request.method == "OPTIONS": return jsonify({}), 200

    try: verify_jwt_in_request()
    except Exception: return jsonify({"error": "Sesión inválida o token ausente"}), 401
    
    operador = get_jwt()
    if operador.get('rol') not in ['SISTEMAS', 'ADMIN', 'LECTURA']:
        return jsonify({"error": "Acceso denegado"}), 403

    id_gen = limpiar_param(request.args.get('generacion'))
    if not id_gen: return jsonify({"error": "Falta Gen"}), 400

    try:
        id_grupo = limpiar_param(request.args.get('grupo')); sem = limpiar_param(request.args.get('semestre')); formato = request.args.get('formato', 'pdf')
        
        try: registrar_accion(operador.get('id'), "DESCARGA_DEUDORES", f"Descargó reporte de deudores (Gen: {id_gen}, Grupo: {id_grupo or 'Todos'}, Sem: {sem or 'Todos'}) en formato {formato.upper()}")
        except: pass

        query = Alumno.query.filter(Alumno.id_generacion == int(id_gen), Alumno.estatus.in_(['ACTIVO', 'BAJA']))
        if id_grupo: query = query.join(Grupo).filter(Grupo.nombre_grupo == id_grupo)
        if sem: query = query.filter(Alumno.semestre_actual == int(sem))
        alumnos = query.all(); alu_dict = {}
        
        hoy = hoy_mx() 
        
        for a in alumnos:
            f_limite = a.fecha_baja if (a.estatus == 'BAJA' and a.fecha_baja) else hoy
            pagos = db.session.query(Pago, EstructuraPago).join(EstructuraPago).filter(Pago.id_alumno == a.id_alumno, Pago.estado.in_(['PENDIENTE', 'PARCIAL'])).all()
            for p, e in pagos:
                if (e.semestre or 1) <= (a.semestre_actual or 1):
                    vencido = False
                    if e.anio and e.mes:
                        if e.anio < f_limite.year or (e.anio == f_limite.year and e.mes <= f_limite.month): vencido = True
                    else: vencido = (a.estatus != 'BAJA')
                    if vencido:
                        deuda = float(e.monto) - float(p.monto_abonado or 0)
                        if deuda > 0:
                            if a.id_alumno not in alu_dict:
                                nom = f"{a.apellido} {a.nombre}".upper() + (" [BAJA]" if a.estatus == 'BAJA' else "")
                                alu_dict[a.id_alumno] = {"Matrícula": a.matricula, "Alumno": nom, "Conceptos": [], "Total": 0}
                            # ASEGURAR PASAR e.semestre AQUI
                            alu_dict[a.id_alumno]["Conceptos"].append({"concepto": e.concepto, "monto": deuda, "semestre": e.semestre})
                            alu_dict[a.id_alumno]["Total"] += deuda
                            
        # APLICAR EL ORDENAMIENTO INTELIGENTE A CADA ALUMNO
        for key in alu_dict:
            alu_dict[key]["Conceptos"] = ordenar_conceptos_escolares(alu_dict[key]["Conceptos"])

        data = list(alu_dict.values())
        
        mensaje_vacio = "No se encontraron alumnos con adeudos pendientes para los filtros seleccionados." 
        titulo = f"ALUMNOS CON DEUDAS" + (f" - {sem}° SEM" if sem else "")

        if formato == 'excel':
            excel_data = [{"Matrícula": x["Matrícula"], "Alumno": x["Alumno"], "Adeudo": x["Total"]} for x in data] if data else []
            return generar_excel(excel_data, "Deudores", "Deudores.xlsx", mensaje_vacio)
        
        return generar_pdf_deudores_pro(data, titulo, "Deudores.pdf", mensaje_vacio)
    except Exception as e: return jsonify({"error": str(e)}), 500


@reportes_bp.route("/reportes/al-corriente", methods=["GET", "OPTIONS"])
def reporte_al_corriente():
    if request.method == "OPTIONS": return jsonify({}), 200

    try: verify_jwt_in_request()
    except Exception: return jsonify({"error": "Sesión inválida o token ausente"}), 401
    
    operador = get_jwt()
    if operador.get('rol') not in ['SISTEMAS', 'ADMIN', 'LECTURA']:
        return jsonify({"error": "Acceso denegado"}), 403

    id_gen = limpiar_param(request.args.get('generacion'))
    try:
        id_grupo = limpiar_param(request.args.get('grupo')); sem = limpiar_param(request.args.get('semestre')); formato = request.args.get('formato', 'pdf')
        
        try: registrar_accion(operador.get('id'), "DESCARGA_AL_CORRIENTE", f"Descargó reporte de alumnos al corriente (Gen: {id_gen}, Grupo: {id_grupo or 'Todos'}, Sem: {sem or 'Todos'})")
        except: pass

        query = Alumno.query.filter(Alumno.id_generacion == int(id_gen), Alumno.estatus == 'ACTIVO')
        if id_grupo: query = query.join(Grupo).filter(Grupo.nombre_grupo == id_grupo)
        if sem: query = query.filter(Alumno.semestre_actual == int(sem))
        
        alumnos = query.all(); data = []
        hoy = hoy_mx()
        
        for a in alumnos:
            tiene_deuda = False
            pagos = db.session.query(Pago, EstructuraPago).join(EstructuraPago).filter(Pago.id_alumno == a.id_alumno, Pago.estado != 'PAGADO').all()
            for p, e in pagos:
                if (e.semestre or 1) <= (a.semestre_actual or 1):
                    if e.anio and e.mes:
                        if e.anio < hoy.year or (e.anio == hoy.year and e.mes <= hoy.month): tiene_deuda = True; break
                    else: tiene_deuda = True; break
            if not tiene_deuda:
                data.append({"Matrícula": a.matricula, "Alumno": f"{a.apellido} {a.nombre}".upper(), "Estatus": "AL CORRIENTE"})
        
        mensaje_vacio = "Actualmente no hay alumnos al corriente en sus pagos bajo los filtros seleccionados." 
        titulo = f"ALUMNOS AL CORRIENTE" + (f" - {sem}° SEM" if sem else "")
        
        if formato == 'excel': 
            return generar_excel(data, "Al Corriente", "Alumnos_Al_Corriente.xlsx", mensaje_vacio)
            
        return generar_pdf_generico_pro(titulo, data, colors.HexColor("#1E293B"), "Al_Corriente.pdf", mensaje_vacio=mensaje_vacio)
    except Exception as e: return jsonify({"error": str(e)}), 500


@reportes_bp.route("/reportes/respaldo", methods=["GET", "OPTIONS"])
def respaldo_db():
    if request.method == "OPTIONS": return jsonify({}), 200
    
    try: verify_jwt_in_request()
    except Exception: return jsonify({"error": "Sesión inválida o token ausente"}), 401
    
    operador = get_jwt()
    if operador.get('rol') != 'SISTEMAS':
        return jsonify({"error": "Solo el nivel SISTEMAS puede generar respaldos físicos de la base de datos"}), 403

    DB_HOST = 'localhost'
    DB_USER = 'root'    
    DB_PASS = ''        
    DB_NAME = 'sistema_prepajmu' 

    ruta_mysqldump = r"C:\xampp\mysql\bin\mysqldump.exe" 
    
    if DB_PASS == '':
        comando = f"{ruta_mysqldump} -h {DB_HOST} -u {DB_USER} {DB_NAME}"
    else:
        comando = f"{ruta_mysqldump} -h {DB_HOST} -u {DB_USER} -p{DB_PASS} {DB_NAME}"

    try:
        proceso = subprocess.Popen(comando, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
        salida, error = proceso.communicate()

        if proceso.returncode != 0:
            print(f" ERROR DE MYSQLDUMP: {error.decode('utf-8', errors='ignore')}")
            return jsonify({"error": "No se pudo crear el respaldo. Revisa la consola."}), 500

        buffer = io.BytesIO(salida)
        buffer.seek(0)
        
        fecha_str = ahora_mx().strftime("%d_%m_%Y_%H%M")
        nombre_archivo = f"Respaldo_Prepa_{fecha_str}.sql"

        try: registrar_accion(operador.get('id'), "RESPALDO_BD", "Descargó un respaldo técnico completo (.sql) de la base de datos.")
        except: pass

        return send_file(buffer, as_attachment=True, download_name=nombre_archivo, mimetype='application/sql')

    except Exception as e: 
        return jsonify({"error": str(e)}), 500
    
@reportes_bp.route("/reportes/adeudo/<matricula>", methods=["GET", "OPTIONS"])
def reporte_adeudo_individual(matricula):
    if request.method == "OPTIONS": return jsonify({}), 200

    # Permitir token por cabecera (Angular HTTP) o por URL (window.open)
    try: verify_jwt_in_request(locations=["headers", "query_string"])
    except Exception: return jsonify({"error": "Sesión inválida o token ausente"}), 401

    operador = get_jwt()
    if operador.get('rol') not in ['SISTEMAS', 'ADMIN', 'LECTURA']:
        return jsonify({"error": "Acceso denegado"}), 403

    try:
        # 1. Buscar al alumno exacto
        alumno = Alumno.query.filter_by(matricula=matricula).first()
        if not alumno:
            return jsonify({"error": "Alumno no encontrado"}), 404

        hoy = hoy_mx()
        f_limite = alumno.fecha_baja if (alumno.estatus == 'BAJA' and alumno.fecha_baja) else hoy

        # 2. Buscar sus pagos pendientes o parciales
        pagos = db.session.query(Pago, EstructuraPago)\
            .join(EstructuraPago)\
            .filter(Pago.id_alumno == alumno.id_alumno, Pago.estado.in_(['PENDIENTE', 'PARCIAL']))\
            .all()

        conceptos_adeudados = []
        total_deuda = 0.0

        for p, e in pagos:
            # Misma lógica estricta de vencimiento que ya usas
            if (e.semestre or 1) <= (alumno.semestre_actual or 1):
                vencido = False
                if e.anio and e.mes:
                    if e.anio < f_limite.year or (e.anio == f_limite.year and e.mes <= f_limite.month): 
                        vencido = True
                else: 
                    vencido = (alumno.estatus != 'BAJA')
                
                if vencido:
                    deuda = float(e.monto) - float(p.monto_abonado or 0)
                    if deuda > 0:
                        # ASEGURAR PASAR e.semestre AQUI
                        conceptos_adeudados.append({"concepto": e.concepto, "monto": deuda, "semestre": e.semestre})
                        total_deuda += deuda
                        
        # APLICAR EL ORDENAMIENTO INTELIGENTE A LA LISTA DEL ALUMNO INDIVIDUAL
        conceptos_adeudados = ordenar_conceptos_escolares(conceptos_adeudados)

        # 3. Formatear para tu motor "generar_pdf_deudores_pro"
        data_agrupada = []
        if total_deuda > 0:
            nombre_str = f"{alumno.apellido} {alumno.nombre}".upper()
            if alumno.estatus == 'BAJA': nombre_str += " [BAJA]"
            
            data_agrupada.append({
                "Matrícula": alumno.matricula,
                "Alumno": nombre_str,
                "Conceptos": conceptos_adeudados,
                "Total": total_deuda
            })

        try: registrar_accion(operador.get('id'), "DESCARGA_AVISO_ADEUDO", f"Descargó aviso de adeudo individual del alumno {matricula}")
        except: pass

        titulo = "AVISO DE PAGO VENCIDO"
        nombre_archivo = f"Aviso_Adeudo_{matricula}.pdf"
        mensaje_vacio = f"El alumno con matrícula {matricula} no tiene adeudos vencidos."

        return generar_pdf_deudores_pro(data_agrupada, titulo, nombre_archivo, mensaje_vacio, descargar=False, es_individual=True)

    except Exception as e: 
        print("ERROR AVISO ADEUDO:", str(e))
        return jsonify({"error": str(e)}), 500